import csv
import io
import re
from collections import Counter
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import redirect_to_login
from django.core.exceptions import PermissionDenied
from django.http import Http404, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from projects.models import Project

from .azure_devops import AzureDevOpsError, fetch_work_items
from .forms import EmployeeForm, EmployeeNoteForm
from .models import AzureDevOpsSettings, Employee, EmployeeNote, EmployeeNoteAttachment, SupportTicket, WorkItem
from .sync import import_team_records, import_team_work_items, run_support_ado_sync, run_team_ado_sync

NOTE_CATEGORY_LABELS = dict(EmployeeNote.Category.choices)


def _clean_ado_item(item):
    """Pop the assignment-matching keys out of a fetch_work_items() dict, leaving only
    valid WorkItem field values (with assigned_to_raw filled in for traceability)."""
    item = dict(item)
    name = item.pop("assigned_to_name", "")
    email = item.pop("assigned_to_email", "")
    item["assigned_to_raw"] = name or email
    return item

MAX_WORK_IMPORT_ROWS = 1000

WORK_IMPORT_FIELDS = [
    {"key": "title", "label": "Title", "required": True},
    {"key": "description", "label": "Description", "required": False},
    {"key": "work_item_type", "label": "Type", "required": False},
    {"key": "state", "label": "State", "required": False},
    {"key": "story_points", "label": "Story Points", "required": False},
    {"key": "area_path", "label": "Area Path", "required": False},
    {"key": "iteration_path", "label": "Iteration Path", "required": False},
    {"key": "created_date", "label": "Created Date", "required": False},
    {"key": "closed_date", "label": "Closed Date", "required": False},
    {"key": "external_id", "label": "External ID", "required": False},
    {"key": "url", "label": "URL", "required": False},
]

ANALYSIS_IMPORT_FIELDS = WORK_IMPORT_FIELDS + [
    {"key": "assignee", "label": "Assignee (name or email)", "required": True},
]


def team_list(request):
    employees = (
        Employee.objects.select_related("manager", "line_manager")
        .prefetch_related("tickets_received", "roles", "projects", "notes")
        .order_by("name")
    )
    rows = []
    for employee in employees:
        notes = list(employee.notes.all())
        rows.append(
            {
                "employee": employee,
                "projects": sorted(employee.projects.all(), key=lambda p: p.name),
                "ticket_count": len(employee.tickets_received.all()),
                "wfh_count": sum(1 for n in notes if n.category == EmployeeNote.Category.WFH_EXCEPTION),
                "achievement_count": sum(1 for n in notes if n.category == EmployeeNote.Category.ACHIEVEMENT),
                "escalation_count": sum(1 for n in notes if n.category == EmployeeNote.Category.ESCALATION),
            }
        )
    return render(request, "teams/team_list.html", {"rows": rows})


def _blended_efficiency(employee):
    values = [v for v in [employee.rtb_efficiency, employee.gsc_efficiency, employee.ai_efficiency] if v is not None]
    if not values:
        return None
    return round(sum(values) / len(values), 2)


def team_efficiency(request):
    employees = Employee.objects.select_related("manager", "line_manager").prefetch_related("work_items")

    groups = {}
    for employee in employees:
        lead = employee.line_manager or employee.manager
        key = lead.pk if lead else 0
        group = groups.setdefault(key, {"manager": lead, "members": []})
        group["members"].append(employee)

    result_groups = []
    for group in groups.values():
        members = sorted(group["members"], key=lambda e: e.name)
        rows = []
        for employee in members:
            items = list(employee.work_items.all())
            rows.append(
                {
                    "employee": employee,
                    "item_count": len(items),
                    "points": sum((i.story_points or 0) for i in items),
                }
            )
        manager = group["manager"]
        result_groups.append(
            {
                "manager": manager,
                "manager_blended": _blended_efficiency(manager) if manager else None,
                "rows": rows,
                "team_size": len(members),
            }
        )

    result_groups.sort(key=lambda g: (g["manager"] is None, g["manager"].name if g["manager"] else ""))
    return render(request, "teams/efficiency.html", {"groups": result_groups})


@login_required
def employee_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == "POST":
        form = EmployeeForm(request.POST, instance=employee, employee=employee)
        if form.is_valid():
            form.save()
            messages.success(request, f"Updated {employee.name}.")
            return redirect("team-org")
    else:
        form = EmployeeForm(instance=employee, employee=employee)
    return render(request, "teams/employee_edit_modal.html", {"employee": employee, "form": form})


def employee_detail(request, pk):
    employee = get_object_or_404(
        Employee.objects.select_related("manager").prefetch_related("reports", "roles", "projects"),
        pk=pk,
    )
    tickets = employee.tickets_received.select_related("project").order_by("-created_at")
    return render(
        request,
        "teams/employee_detail.html",
        {"employee": employee, "tickets": tickets, "projects": employee.projects.all()},
    )


@login_required
def employee_projects_edit(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    all_projects = Project.objects.order_by("name")

    if request.method == "POST":
        selected_ids = request.POST.getlist("projects")
        employee.projects.set(all_projects.filter(pk__in=selected_ids))
        messages.success(request, f"Updated projects for {employee.name}.")
        return redirect("team-org")

    current_ids = set(employee.projects.values_list("pk", flat=True))
    return render(
        request,
        "teams/projects_edit.html",
        {"employee": employee, "all_projects": all_projects, "current_ids": current_ids},
    )


def _analysis_for(employee, items):
    monthly_counts = Counter()
    for item in items:
        d = item.closed_date or item.created_date
        if d:
            monthly_counts[d.strftime("%Y-%m")] += 1
    max_count = max(monthly_counts.values(), default=0)
    monthly_series = [
        {"month": month, "count": count, "pct": round(count / max_count * 100) if max_count else 0}
        for month, count in sorted(monthly_counts.items())
    ]

    type_counts = Counter(i.work_item_type or "Unspecified" for i in items).most_common()
    state_counts = Counter(i.state or "Unspecified" for i in items).most_common()
    total_points = sum((i.story_points or 0) for i in items)

    return {
        "items": items[:50],
        "total_items": len(items),
        "monthly_series": monthly_series,
        "type_counts": type_counts,
        "state_counts": state_counts,
        "total_points": total_points,
        "prompt": _build_analysis_prompt(employee, items, monthly_series, type_counts, total_points),
    }


def _build_analysis_prompt(employee, items, monthly_series, type_counts, total_points):
    if not items:
        return (
            f"No work items are on file yet for {employee.name}. Sync from Azure DevOps or import "
            "an Excel/CSV file for this person, then reopen this analysis."
        )

    dated = [i.closed_date or i.created_date for i in items if (i.closed_date or i.created_date)]
    span = f" from {min(dated):%b %Y} to {max(dated):%b %Y}" if dated else ""
    type_lines = "\n".join(f"- {t}: {c}" for t, c in type_counts)
    monthly_lines = "\n".join(f"- {m['month']}: {m['count']} item(s)" for m in monthly_series)
    sample_titles = "\n".join(f"- [{i.work_item_type or 'Item'}] {i.title}" for i in items[:25])

    roles = employee.roles_display() or "role unspecified"
    return (
        f"Analyze {employee.name}'s ({roles}) delivery history{span}, "
        f"based on {len(items)} tracked work item(s) totalling {total_points or 0} story points.\n\n"
        f"Work item types:\n{type_lines}\n\n"
        f"Monthly delivery:\n{monthly_lines}\n\n"
        f"Sample tickets:\n{sample_titles}\n\n"
        "Please summarize: (1) the kinds of tickets/user stories this person typically works on, "
        "(2) an effort analysis (volume and story points per month, and any trend), and "
        "(3) an overall narrative summary of this developer's contribution since the start of the project."
    )


def employee_work(request, pk):
    employee = get_object_or_404(Employee.objects.prefetch_related("roles"), pk=pk)
    settings_obj = AzureDevOpsSettings.load()
    ado_error = None

    default_source = (
        "azure_devops" if (settings_obj.personal_access_token and employee.azure_devops_query_url) else "excel"
    )
    source = request.GET.get("source") or default_source
    if source not in ("azure_devops", "excel"):
        source = default_source

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())

        if request.POST.get("action") == "ado_sync":
            source = "azure_devops"
            query_url = request.POST.get("query_url", "").strip()
            employee.azure_devops_query_url = query_url
            employee.save(update_fields=["azure_devops_query_url"])

            if not query_url:
                ado_error = "Add a query URL first."
            elif not settings_obj.personal_access_token:
                ado_error = "No shared PAT is configured. Ask an admin to set one under Admin > Teams > Azure DevOps settings."
            else:
                try:
                    fetched = fetch_work_items(query_url, settings_obj.personal_access_token)
                except AzureDevOpsError as exc:
                    ado_error = str(exc)
                else:
                    employee.work_items.filter(source=WorkItem.Source.AZURE_DEVOPS).delete()
                    WorkItem.objects.bulk_create(
                        WorkItem(employee=employee, source=WorkItem.Source.AZURE_DEVOPS, **_clean_ado_item(item))
                        for item in fetched
                    )
                    messages.success(request, f"Synced {len(fetched)} work item(s) from Azure DevOps.")
                    return redirect(f"{reverse('employee-work', args=[pk])}?source=azure_devops")

    source_value = WorkItem.Source.AZURE_DEVOPS if source == "azure_devops" else WorkItem.Source.EXCEL
    items = list(employee.work_items.filter(source=source_value))

    context = {
        "employee": employee,
        "source": source,
        "ado_error": ado_error,
        "pat_configured": bool(settings_obj.personal_access_token),
        **_analysis_for(employee, items),
    }
    return render(request, "teams/work.html", context)


def _read_tabular_file(upload):
    name = (upload.name or "").lower()
    if name.endswith(".csv"):
        text = io.TextIOWrapper(upload.file, encoding="utf-8-sig", errors="replace")
        rows = [row for row in csv.reader(text)]
        if not rows:
            raise ValueError("That CSV file is empty.")
        headers = [cell.strip() for cell in rows[0]]
        data_rows = [[cell.strip() for cell in row] for row in rows[1:] if any(cell.strip() for cell in row)]
        return headers, data_rows

    try:
        workbook = openpyxl.load_workbook(upload, read_only=True, data_only=True)
        worksheet = workbook.active
        rows_iter = worksheet.iter_rows(values_only=True)
        header_row = next(rows_iter)
    except Exception as exc:
        raise ValueError("Could not read that file. Please upload a valid .xlsx or .csv file.") from exc

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]
    data_rows = []
    for row in rows_iter:
        if all(cell is None for cell in row):
            continue
        data_rows.append(["" if cell is None else str(cell).strip() for cell in row])
    return headers, data_rows


def _guess_work_column(headers, key):
    key_variants = {key.lower(), key.replace("_", " ").lower()}
    for header in headers:
        h = header.lower()
        if any(variant in h for variant in key_variants):
            return header
    return ""


@login_required
def employee_import_upload(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Choose an .xlsx or .csv file to upload.")
            return redirect("employee-import-upload", pk=pk)

        try:
            headers, data_rows = _read_tabular_file(upload)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("employee-import-upload", pk=pk)

        if not any(headers) or not data_rows:
            messages.error(request, "That file needs a header row and at least one data row.")
            return redirect("employee-import-upload", pk=pk)

        request.session["work_import_headers"] = headers
        request.session["work_import_rows"] = data_rows[:MAX_WORK_IMPORT_ROWS]
        return redirect("employee-import-map", pk=pk)

    return render(request, "teams/work_import_upload.html", {"employee": employee})


def _parse_work_date(value):
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_story_points(value):
    value = (value or "").strip()
    if not value:
        return None
    try:
        return Decimal(value)
    except (InvalidOperation, ValueError):
        return None


@login_required
def employee_import_map(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    headers = request.session.get("work_import_headers")
    rows = request.session.get("work_import_rows")
    if not headers or not rows:
        messages.error(request, "Upload a file first.")
        return redirect("employee-import-upload", pk=pk)

    if request.method == "POST":
        mapping = {f["key"]: request.POST.get(f"map_{f['key']}", "") for f in WORK_IMPORT_FIELDS}
        missing_required = [f["label"] for f in WORK_IMPORT_FIELDS if f["required"] and not mapping[f["key"]]]
        if missing_required:
            messages.error(request, f"Map a column to: {', '.join(missing_required)}.")
            return render(
                request,
                "teams/work_import_map.html",
                {"employee": employee, "headers": headers, "rows": rows[:8], "fields": WORK_IMPORT_FIELDS, "mapping": mapping},
            )

        col_index = {header: i for i, header in enumerate(headers)}

        def cell(row, key):
            header = mapping[key]
            idx = col_index.get(header)
            if not header or idx is None or idx >= len(row):
                return ""
            return row[idx]

        created = 0
        for row in rows:
            title = cell(row, "title")
            if not title:
                continue
            WorkItem.objects.create(
                employee=employee,
                source=WorkItem.Source.EXCEL,
                title=title,
                description=cell(row, "description"),
                work_item_type=cell(row, "work_item_type"),
                state=cell(row, "state"),
                story_points=_parse_story_points(cell(row, "story_points")),
                area_path=cell(row, "area_path"),
                iteration_path=cell(row, "iteration_path"),
                created_date=_parse_work_date(cell(row, "created_date")),
                closed_date=_parse_work_date(cell(row, "closed_date")),
                external_id=cell(row, "external_id"),
                url=cell(row, "url"),
            )
            created += 1

        del request.session["work_import_headers"]
        del request.session["work_import_rows"]
        messages.success(request, f"Imported {created} work item(s) for {employee.name}.")
        return redirect(f"{reverse('employee-work', args=[pk])}?source=excel")

    mapping = {f["key"]: _guess_work_column(headers, f["key"]) for f in WORK_IMPORT_FIELDS}
    return render(
        request,
        "teams/work_import_map.html",
        {"employee": employee, "headers": headers, "rows": rows[:8], "fields": WORK_IMPORT_FIELDS, "mapping": mapping},
    )


def _notes_list_context(employee, category, form=None):
    return {
        "employee": employee,
        "category": category,
        "category_label": NOTE_CATEGORY_LABELS[category],
        "notes": employee.notes.filter(category=category),
        "form": form or EmployeeNoteForm(),
    }


def employee_notes(request, pk, category):
    employee = get_object_or_404(Employee, pk=pk)
    if category not in NOTE_CATEGORY_LABELS:
        raise Http404("Unknown note category.")

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        form = EmployeeNoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.employee = employee
            note.category = category
            note.save()
            return render(request, "teams/_notes_list.html", _notes_list_context(employee, category))
        return render(request, "teams/_notes_list.html", _notes_list_context(employee, category, form))

    return render(request, "teams/_notes_list.html", _notes_list_context(employee, category))


def employee_note_edit(request, pk, category, note_id):
    employee = get_object_or_404(Employee, pk=pk)
    if category not in NOTE_CATEGORY_LABELS:
        raise Http404("Unknown note category.")
    note = get_object_or_404(EmployeeNote, pk=note_id, employee=employee, category=category)

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        form = EmployeeNoteForm(request.POST, instance=note)
        if form.is_valid():
            form.save()
            return render(request, "teams/_notes_list.html", _notes_list_context(employee, category))
    else:
        form = EmployeeNoteForm(instance=note)

    return render(
        request,
        "teams/_note_edit_form.html",
        {
            "employee": employee,
            "category": category,
            "category_label": NOTE_CATEGORY_LABELS[category],
            "note": note,
            "form": form,
        },
    )


def employee_note_delete(request, pk, category, note_id):
    employee = get_object_or_404(Employee, pk=pk)
    if category not in NOTE_CATEGORY_LABELS:
        raise Http404("Unknown note category.")
    note = get_object_or_404(EmployeeNote, pk=note_id, employee=employee, category=category)

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        note.delete()
        return render(request, "teams/_notes_list.html", _notes_list_context(employee, category))

    return redirect("employee-notes", pk=pk, category=category)


def employee_note_attachment_add(request, pk, category, note_id):
    employee = get_object_or_404(Employee, pk=pk)
    if category not in NOTE_CATEGORY_LABELS:
        raise Http404("Unknown note category.")
    note = get_object_or_404(EmployeeNote, pk=note_id, employee=employee, category=category)

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        for uploaded_file in request.FILES.getlist("file"):
            EmployeeNoteAttachment.objects.create(note=note, file=uploaded_file)
        return render(request, "teams/_notes_list.html", _notes_list_context(employee, category))

    return redirect("employee-notes", pk=pk, category=category)


def employee_note_attachment_delete(request, pk, category, note_id, attachment_id):
    employee = get_object_or_404(Employee, pk=pk)
    if category not in NOTE_CATEGORY_LABELS:
        raise Http404("Unknown note category.")
    note = get_object_or_404(EmployeeNote, pk=note_id, employee=employee, category=category)
    attachment = get_object_or_404(EmployeeNoteAttachment, pk=attachment_id, note=note)

    if request.method == "POST":
        if not request.user.is_authenticated:
            return redirect_to_login(request.get_full_path())
        attachment.file.delete(save=False)
        attachment.delete()
        return render(request, "teams/_notes_list.html", _notes_list_context(employee, category))

    return redirect("employee-notes", pk=pk, category=category)


def _flash_import_result(request, result):
    if result["matched"] == 0 and not result["unmatched"]:
        messages.warning(request, "No user stories were found.")
        return
    base = f"Imported {result['matched']} item(s) across {result['employee_count']} employee(s)."
    if result["unmatched"]:
        preview = "; ".join(result["unmatched"][:8])
        more = f" (+{len(result['unmatched']) - 8} more)" if len(result["unmatched"]) > 8 else ""
        messages.warning(
            request,
            f"{base} {len(result['unmatched'])} item(s) skipped — no employee matched: {preview}{more}",
        )
    else:
        messages.success(request, base)


MONTH_RANGE_OPTIONS = list(range(1, 13))

AGING_BUCKET_LABELS = ["0–7 days", "8–14 days", "15–30 days", "31+ days"]


def _bar_series(counter, link_for=None):
    max_count = max(counter.values(), default=0)
    rows = []
    for label, count in counter.most_common():
        row = {"label": label or "Unspecified", "count": count, "pct": round(count / max_count * 100) if max_count else 0}
        if link_for:
            href = link_for(label)
            if href:
                row["href"] = href
        rows.append(row)
    return rows


def _split_tags(raw):
    if not raw:
        return []
    return [part.strip() for part in re.split(r"[;,]", raw) if part.strip()]


def _filter_link(request, param, value):
    """Build a URL for the current page that keeps every other active filter but
    narrows `param` down to a single `value` — used to make chart bars and heatmap
    row labels clickable drill-down filters."""
    params = request.GET.copy()
    params.setlist(param, [str(value)])
    return "?" + params.urlencode()


def _heat_level(count, max_count):
    if not count:
        return 0
    if not max_count:
        return 0
    pct = count / max_count
    if pct >= 0.75:
        return 4
    if pct >= 0.5:
        return 3
    if pct >= 0.25:
        return 2
    return 1


def _cross_matrix(counts_by_row, columns, hrefs=None):
    """Build a row-label x column-label count matrix (used for month columns and,
    separately, country columns), each cell shaded by a heat level relative to the
    matrix's own max cell value. `hrefs`, if given, is a {label: url} map used to
    make row labels clickable drill-down filters."""
    all_counts = [c for row in counts_by_row.values() for c in row.values()]
    max_count = max(all_counts, default=0)
    rows = []
    for label, col_counts in counts_by_row.items():
        cells = [
            {"count": col_counts.get(col, 0), "level": _heat_level(col_counts.get(col, 0), max_count)}
            for col in columns
        ]
        row = {"label": label, "total": sum(col_counts.values()), "cells": cells}
        if hrefs is not None:
            row["href"] = hrefs.get(label)
        rows.append(row)
    rows.sort(key=lambda r: -r["total"])
    return rows


def _parse_month_filter(request):
    raw = request.GET.get("months", "").strip()
    try:
        months = int(raw)
    except ValueError:
        return None
    return months if months in MONTH_RANGE_OPTIONS else None


def _months_cutoff(months):
    if not months:
        return None
    today = timezone.localdate()
    year, month = today.year, today.month - (months - 1)
    while month <= 0:
        month += 12
        year -= 1
    return date(year, month, 1)


def _team_ticket_analysis_data(request, model):
    """Filter `model` records (WorkItem or SupportTicket) by the months/developer/
    country query params and build every aggregate the page needs (bar charts,
    monthly heatmaps, prompt) from that single filtered set, so everything on the
    page reflects the same selection."""
    months = _parse_month_filter(request)
    cutoff = _months_cutoff(months)
    developer_ids = request.GET.getlist("developer")
    countries = request.GET.getlist("country")
    query = request.GET.get("q", "").strip()

    employee_ids_with_data = model.objects.values_list("employee_id", flat=True).distinct()
    filter_employees = list(Employee.objects.filter(pk__in=employee_ids_with_data).order_by("name"))
    filter_countries = sorted({e.country or "Unspecified" for e in filter_employees})
    name_to_pk = {e.name: e.pk for e in filter_employees}

    allowed_employee_pks = {
        e.pk
        for e in filter_employees
        if (not developer_ids or str(e.pk) in developer_ids)
        and (not countries or (e.country or "Unspecified") in countries)
    }

    related_select = ["employee"] + (["related_work_item"] if hasattr(model, "related_work_item") else [])
    all_items = model.objects.filter(employee_id__in=allowed_employee_pks).select_related(*related_select)
    items = []
    for item in all_items:
        d = item.closed_date or item.created_date
        if cutoff and (not d or d < cutoff):
            continue
        if query and query.lower() not in f"{item.title} {item.description}".lower():
            continue
        items.append(item)

    by_developer_count = Counter()
    by_country_count = Counter()
    by_type = Counter()
    by_component_count = Counter()
    by_priority_count = Counter()
    by_tag_count = Counter()
    monthly = Counter()
    dev_month_counts = {}
    country_month_counts = {}
    component_country_counts = {}
    rows_by_employee = {}
    countries_present = set()
    total_points = 0
    resolution_days = []
    open_buckets = Counter()
    open_no_date = 0
    today = timezone.localdate()

    for item in items:
        employee = item.employee
        emp_country = employee.country or "Unspecified"
        component = item.area_path or "Unspecified"
        countries_present.add(emp_country)

        by_developer_count[employee.name] += 1
        by_country_count[emp_country] += 1
        by_type[item.work_item_type or "Unspecified"] += 1
        by_component_count[component] += 1
        by_priority_count[f"Priority {item.priority}" if item.priority else "Unspecified"] += 1
        for tag in _split_tags(item.tags):
            by_tag_count[tag] += 1
        component_country_counts.setdefault(component, Counter())[emp_country] += 1

        points = item.story_points or 0
        total_points += points

        row = rows_by_employee.setdefault(employee.pk, {"employee": employee, "item_count": 0, "points": 0})
        row["item_count"] += 1
        row["points"] += points

        if item.closed_date and item.created_date:
            resolution_days.append((item.closed_date - item.created_date).days)
        if not item.closed_date:
            basis = item.created_date
            if basis:
                age = (today - basis).days
                if age <= 7:
                    open_buckets[AGING_BUCKET_LABELS[0]] += 1
                elif age <= 14:
                    open_buckets[AGING_BUCKET_LABELS[1]] += 1
                elif age <= 30:
                    open_buckets[AGING_BUCKET_LABELS[2]] += 1
                else:
                    open_buckets[AGING_BUCKET_LABELS[3]] += 1
            else:
                open_no_date += 1

        d = item.closed_date or item.created_date
        if d:
            month = d.strftime("%Y-%m")
            monthly[month] += 1
            dev_month_counts.setdefault(employee.name, Counter())[month] += 1
            country_month_counts.setdefault(emp_country, Counter())[month] += 1

    months_columns = sorted(monthly.keys())
    countries_columns = sorted(countries_present)
    max_monthly = max(monthly.values(), default=0)
    monthly_sorted = sorted(monthly.items())
    monthly_series = [
        {"month": month, "count": count, "pct": round(count / max_monthly * 100) if max_monthly else 0}
        for month, count in monthly_sorted
    ]

    trend = None
    if len(monthly_sorted) >= 2:
        prev_month, prev_count = monthly_sorted[-2]
        last_month, last_count = monthly_sorted[-1]
        if prev_count:
            pct = round((last_count - prev_count) / prev_count * 100)
        else:
            pct = 100 if last_count else 0
        trend = {
            "prev_month": prev_month,
            "prev_count": prev_count,
            "last_month": last_month,
            "last_count": last_count,
            "pct": pct,
            "direction": "up" if last_count > prev_count else ("down" if last_count < prev_count else "flat"),
        }

    max_bucket = max(open_buckets.values(), default=0)
    aging = {
        "avg_resolution_days": round(sum(resolution_days) / len(resolution_days), 1) if resolution_days else None,
        "resolved_count": len(resolution_days),
        "open_count": sum(open_buckets.values()) + open_no_date,
        "open_no_date": open_no_date,
        "open_buckets": [
            {
                "label": label,
                "count": open_buckets.get(label, 0),
                "pct": round(open_buckets.get(label, 0) / max_bucket * 100) if max_bucket else 0,
            }
            for label in AGING_BUCKET_LABELS
        ],
    }

    summary_rows = sorted(rows_by_employee.values(), key=lambda r: -r["item_count"])

    story_rows = sorted(items, key=lambda i: i.closed_date or i.created_date or date.min, reverse=True)
    story_row_limit = 150

    dev_link = lambda name: _filter_link(request, "developer", name_to_pk[name]) if name in name_to_pk else None
    country_link = lambda name: _filter_link(request, "country", name)
    developer_hrefs = {name: _filter_link(request, "developer", pk) for name, pk in name_to_pk.items()}
    country_hrefs = {c: _filter_link(request, "country", c) for c in countries_present}

    return {
        "items": items,
        "summary": {
            "rows": summary_rows,
            "total_items": len(items),
            "total_points": total_points,
            "employee_count": len(summary_rows),
            "trend": trend,
        },
        "aging": aging,
        "charts": {
            "by_developer": _bar_series(by_developer_count, link_for=dev_link),
            "by_country": _bar_series(by_country_count, link_for=country_link),
            "by_type": by_type.most_common(),
            "by_component": _bar_series(by_component_count),
            "by_priority": _bar_series(by_priority_count),
            "by_tag": _bar_series(by_tag_count)[:12],
            "monthly_series": monthly_series,
        },
        "months_columns": months_columns,
        "developer_matrix": _cross_matrix(dev_month_counts, months_columns, hrefs=developer_hrefs),
        "country_matrix": _cross_matrix(country_month_counts, months_columns, hrefs=country_hrefs),
        "countries_columns": countries_columns,
        "component_matrix": _cross_matrix(component_country_counts, countries_columns),
        "story_rows": [
            {
                "employee": i.employee,
                "title": i.title,
                "description": i.description,
                "country": i.employee.country or "Unspecified",
                "type": i.work_item_type,
                "component": i.area_path,
                "priority": i.priority,
                "tags": i.tags,
                "url": i.url,
                "related_work_item": getattr(i, "related_work_item", None),
            }
            for i in story_rows[:story_row_limit]
        ],
        "story_row_total": len(story_rows),
        "story_row_limit": story_row_limit,
        "filters": {
            "month_options": MONTH_RANGE_OPTIONS,
            "selected_months": months,
            "filter_employees": filter_employees,
            "selected_developer_ids": developer_ids,
            "filter_countries": filter_countries,
            "selected_countries": countries,
            "query": query,
            "active": bool(months or developer_ids or countries or query),
        },
    }


def _build_ticket_analysis_prompt(data, record_noun="user stories/tickets", config_tab_name="Configuration"):
    summary = data["summary"]
    charts = data["charts"]
    total_items = summary["total_items"]
    if not total_items:
        return (
            f"No {record_noun} are on file yet for the current filter. Sync from Azure DevOps or "
            f"import an Excel/CSV file on the {config_tab_name} tab, then reopen this prompt."
        )

    filters = data["filters"]
    scope_bits = []
    if filters["selected_months"]:
        scope_bits.append(f"the last {filters['selected_months']} month(s)")
    if filters["selected_developer_ids"]:
        names = [e.name for e in filters["filter_employees"] if str(e.pk) in filters["selected_developer_ids"]]
        scope_bits.append(f"developers: {', '.join(names)}")
    if filters["selected_countries"]:
        scope_bits.append(f"countries: {', '.join(filters['selected_countries'])}")
    scope_line = f"Scope: {'; '.join(scope_bits)}.\n\n" if scope_bits else ""

    dev_lines = "\n".join(f"- {d['label']}: {d['count']}" for d in charts["by_developer"])
    country_lines = "\n".join(f"- {c['label']}: {c['count']}" for c in charts["by_country"])
    type_lines = "\n".join(f"- {t}: {c}" for t, c in charts["by_type"])
    component_lines = "\n".join(f"- {c['label']}: {c['count']}" for c in charts["by_component"])
    priority_lines = "\n".join(f"- {p['label']}: {p['count']}" for p in charts["by_priority"]) or "- (none set)"
    tag_lines = "\n".join(f"- {t['label']}: {t['count']}" for t in charts["by_tag"]) or "- (none set)"
    monthly_lines = "\n".join(f"- {m['month']}: {m['count']} item(s)" for m in charts["monthly_series"])

    aging = data.get("aging") or {}
    aging_line = ""
    if aging.get("avg_resolution_days") is not None:
        aging_line = (
            f"\nAverage resolution time: {aging['avg_resolution_days']} day(s) across "
            f"{aging['resolved_count']} closed item(s). Currently open: {aging['open_count']}.\n"
        )

    sample_lines = []
    for i in data["items"][:60]:
        line = f"- [{i.work_item_type or 'Item'}] {i.title}"
        if i.area_path:
            line += f" (Area Path: {i.area_path})"
        if i.description:
            line += f" — {i.description[:200]}"
        sample_lines.append(line)
    sample_titles = "\n".join(sample_lines)
    more_note = ""
    if total_items > 60:
        more_note = f" (showing 60 of {total_items} — ask for more if you need the rest)"

    return (
        f"{scope_line}"
        f"Analyze this team's delivery history based on {total_items} tracked {record_noun} "
        f"totalling {summary['total_points'] or 0} story points.\n\n"
        f"Reference data (from Azure DevOps fields, for context only — please form your own "
        f"judgment from the ticket text below rather than just repeating these numbers):\n"
        f"By developer:\n{dev_lines}\n\n"
        f"By country:\n{country_lines}\n\n"
        f"By work item type:\n{type_lines}\n\n"
        f"By Area Path (as tagged in the source system, may be incomplete or inconsistently used):\n{component_lines}\n\n"
        f"By priority:\n{priority_lines}\n\n"
        f"By tag:\n{tag_lines}\n\n"
        f"Monthly delivery trend:\n{monthly_lines}\n"
        f"{aging_line}\n"
        f"Tickets{more_note} — read the title and description of each to do your own analysis:\n{sample_titles}\n\n"
        "Based on reading the actual ticket titles and descriptions above (not just the reference "
        "counts), please: (1) classify each ticket, or group of similar tickets, as primarily "
        "'configuration' work (settings, environment, access, parameter changes) or 'development' "
        "work (new features, bug fixes, code changes), and give an overall estimate of the split "
        "with your reasoning, (2) identify the actual products, components, or subsystems these "
        "tickets relate to based on their content — cross-check against the Area Path data above but "
        "correct or refine it where the ticket text suggests a better grouping, especially for any "
        "'Unspecified' ones, (3) name the request types/themes that come up a lot, (4) describe the "
        "overall delivery trend over time and whether it's accelerating, steady, or slowing, (5) note "
        "how work is distributed across countries/locations and developers, and whether that looks "
        "balanced, and (6) summarize the important subjects/themes this team has been working on."
    )


def _render_ticket_analysis(request, model, template, record_noun):
    data = _team_ticket_analysis_data(request, model)
    return render(
        request,
        template,
        {
            "summary": data["summary"],
            "charts": data["charts"],
            "months_columns": data["months_columns"],
            "developer_matrix": data["developer_matrix"],
            "country_matrix": data["country_matrix"],
            "countries_columns": data["countries_columns"],
            "component_matrix": data["component_matrix"],
            "story_rows": data["story_rows"],
            "story_row_total": data["story_row_total"],
            "story_row_limit": data["story_row_limit"],
            "filters": data["filters"],
            "aging": data["aging"],
            "prompt": _build_ticket_analysis_prompt(data, record_noun=record_noun),
        },
    )


def _export_tickets_excel(request, model, filename_prefix):
    data = _team_ticket_analysis_data(request, model)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename_prefix[:31]
    ws.append([
        "Owner", "Title", "Description", "Type", "State", "Component", "Country",
        "Priority", "Tags", "Points/Effort", "Created", "Closed", "URL", "Linked Change",
    ])
    for item in data["items"]:
        related = getattr(item, "related_work_item", None)
        ws.append([
            item.employee.name,
            item.title,
            item.description,
            item.work_item_type,
            item.state,
            item.area_path,
            item.employee.country or "Unspecified",
            item.priority,
            item.tags,
            float(item.story_points) if item.story_points is not None else None,
            item.created_date.isoformat() if item.created_date else "",
            item.closed_date.isoformat() if item.closed_date else "",
            item.url,
            related.title if related else "",
        ])
    buf = io.BytesIO()
    wb.save(buf)
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename_prefix}.xlsx"'
    return response


def analysis_export(request):
    return _export_tickets_excel(request, WorkItem, "change_management_export")


def support_export(request):
    return _export_tickets_excel(request, SupportTicket, "support_tickets_export")


def analysis_home(request):
    settings_obj = AzureDevOpsSettings.load()

    due = (
        settings_obj.last_synced_at is None
        or timezone.now() - settings_obj.last_synced_at >= timedelta(hours=settings_obj.auto_sync_interval_hours)
    )
    if settings_obj.auto_sync_enabled and settings_obj.team_query_url and settings_obj.personal_access_token and due:
        run_team_ado_sync(settings_obj)

    return _render_ticket_analysis(request, WorkItem, "teams/analysis.html", "user stories/tickets")


def support_analysis_home(request):
    settings_obj = AzureDevOpsSettings.load()

    due = (
        settings_obj.support_last_synced_at is None
        or timezone.now() - settings_obj.support_last_synced_at
        >= timedelta(hours=settings_obj.support_auto_sync_interval_hours)
    )
    if (
        settings_obj.support_auto_sync_enabled
        and settings_obj.support_query_url
        and settings_obj.personal_access_token
        and due
    ):
        run_support_ado_sync(settings_obj)

    return _render_ticket_analysis(request, SupportTicket, "teams/support_analysis.html", "support tickets")


def admin_settings(request):
    if not request.user.is_authenticated:
        return redirect_to_login(request.get_full_path())
    if not request.user.is_staff:
        raise PermissionDenied("Admin settings are staff-only.")

    settings_obj = AzureDevOpsSettings.load()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "sync_now":
            result, error = run_team_ado_sync(settings_obj)
            if error:
                messages.error(request, f"Sync failed: {error}")
            else:
                _flash_import_result(request, result)
            return redirect("admin-settings")

        if action == "sync_now_support":
            result, error = run_support_ado_sync(settings_obj)
            if error:
                messages.error(request, f"Sync failed: {error}")
            else:
                _flash_import_result(request, result)
            return redirect("admin-settings")

        settings_obj.organization_url = request.POST.get("organization_url", "").strip()
        settings_obj.personal_access_token = request.POST.get("personal_access_token", "").strip()
        settings_obj.team_query_url = request.POST.get("team_query_url", "").strip()
        settings_obj.auto_sync_enabled = bool(request.POST.get("auto_sync_enabled"))
        try:
            interval = int(request.POST.get("auto_sync_interval_hours", "24"))
        except ValueError:
            interval = 24
        settings_obj.auto_sync_interval_hours = max(1, interval)

        settings_obj.support_query_url = request.POST.get("support_query_url", "").strip()
        settings_obj.support_auto_sync_enabled = bool(request.POST.get("support_auto_sync_enabled"))
        try:
            support_interval = int(request.POST.get("support_auto_sync_interval_hours", "24"))
        except ValueError:
            support_interval = 24
        settings_obj.support_auto_sync_interval_hours = max(1, support_interval)

        settings_obj.save()
        messages.success(request, "Settings saved.")
        return redirect("admin-settings")

    return render(request, "teams/admin_settings.html", {"settings_obj": settings_obj})


@login_required
def analysis_import_upload(request):
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Choose an .xlsx or .csv file to upload.")
            return redirect("analysis-import-upload")

        try:
            headers, data_rows = _read_tabular_file(upload)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("analysis-import-upload")

        if not any(headers) or not data_rows:
            messages.error(request, "That file needs a header row and at least one data row.")
            return redirect("analysis-import-upload")

        request.session["analysis_import_headers"] = headers
        request.session["analysis_import_rows"] = data_rows[:MAX_WORK_IMPORT_ROWS]
        return redirect("analysis-import-map")

    return render(request, "teams/analysis_import_upload.html")


@login_required
def analysis_import_map(request):
    headers = request.session.get("analysis_import_headers")
    rows = request.session.get("analysis_import_rows")
    if not headers or not rows:
        messages.error(request, "Upload a file first.")
        return redirect("analysis-import-upload")

    if request.method == "POST":
        mapping = {f["key"]: request.POST.get(f"map_{f['key']}", "") for f in ANALYSIS_IMPORT_FIELDS}
        missing_required = [f["label"] for f in ANALYSIS_IMPORT_FIELDS if f["required"] and not mapping[f["key"]]]
        if missing_required:
            messages.error(request, f"Map a column to: {', '.join(missing_required)}.")
            return render(
                request,
                "teams/analysis_import_map.html",
                {"headers": headers, "rows": rows[:8], "fields": ANALYSIS_IMPORT_FIELDS, "mapping": mapping},
            )

        col_index = {header: i for i, header in enumerate(headers)}

        def cell(row, key):
            header = mapping[key]
            idx = col_index.get(header)
            if not header or idx is None or idx >= len(row):
                return ""
            return row[idx]

        items = []
        for row in rows:
            title = cell(row, "title")
            if not title:
                continue
            assignee = cell(row, "assignee")
            items.append(
                {
                    "title": title,
                    "description": cell(row, "description"),
                    "work_item_type": cell(row, "work_item_type"),
                    "state": cell(row, "state"),
                    "story_points": _parse_story_points(cell(row, "story_points")),
                    "area_path": cell(row, "area_path"),
                    "iteration_path": cell(row, "iteration_path"),
                    "created_date": _parse_work_date(cell(row, "created_date")),
                    "closed_date": _parse_work_date(cell(row, "closed_date")),
                    "external_id": cell(row, "external_id"),
                    "url": cell(row, "url"),
                    "assigned_to_name": assignee,
                    "assigned_to_email": assignee,
                }
            )

        result = import_team_work_items(items, source=WorkItem.Source.EXCEL, upsert=False)
        del request.session["analysis_import_headers"]
        del request.session["analysis_import_rows"]
        _flash_import_result(request, result)
        return redirect("team-analysis")

    mapping = {f["key"]: _guess_work_column(headers, f["key"]) for f in ANALYSIS_IMPORT_FIELDS}
    return render(
        request,
        "teams/analysis_import_map.html",
        {"headers": headers, "rows": rows[:8], "fields": ANALYSIS_IMPORT_FIELDS, "mapping": mapping},
    )


@login_required
def support_import_upload(request):
    if request.method == "POST":
        upload = request.FILES.get("file")
        if not upload:
            messages.error(request, "Choose an .xlsx or .csv file to upload.")
            return redirect("support-import-upload")

        try:
            headers, data_rows = _read_tabular_file(upload)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("support-import-upload")

        if not any(headers) or not data_rows:
            messages.error(request, "That file needs a header row and at least one data row.")
            return redirect("support-import-upload")

        request.session["support_import_headers"] = headers
        request.session["support_import_rows"] = data_rows[:MAX_WORK_IMPORT_ROWS]
        return redirect("support-import-map")

    return render(request, "teams/support_import_upload.html")


@login_required
def support_import_map(request):
    headers = request.session.get("support_import_headers")
    rows = request.session.get("support_import_rows")
    if not headers or not rows:
        messages.error(request, "Upload a file first.")
        return redirect("support-import-upload")

    if request.method == "POST":
        mapping = {f["key"]: request.POST.get(f"map_{f['key']}", "") for f in ANALYSIS_IMPORT_FIELDS}
        missing_required = [f["label"] for f in ANALYSIS_IMPORT_FIELDS if f["required"] and not mapping[f["key"]]]
        if missing_required:
            messages.error(request, f"Map a column to: {', '.join(missing_required)}.")
            return render(
                request,
                "teams/support_import_map.html",
                {"headers": headers, "rows": rows[:8], "fields": ANALYSIS_IMPORT_FIELDS, "mapping": mapping},
            )

        col_index = {header: i for i, header in enumerate(headers)}

        def cell(row, key):
            header = mapping[key]
            idx = col_index.get(header)
            if not header or idx is None or idx >= len(row):
                return ""
            return row[idx]

        items = []
        for row in rows:
            title = cell(row, "title")
            if not title:
                continue
            assignee = cell(row, "assignee")
            items.append(
                {
                    "title": title,
                    "description": cell(row, "description"),
                    "work_item_type": cell(row, "work_item_type"),
                    "state": cell(row, "state"),
                    "story_points": _parse_story_points(cell(row, "story_points")),
                    "area_path": cell(row, "area_path"),
                    "iteration_path": cell(row, "iteration_path"),
                    "created_date": _parse_work_date(cell(row, "created_date")),
                    "closed_date": _parse_work_date(cell(row, "closed_date")),
                    "external_id": cell(row, "external_id"),
                    "url": cell(row, "url"),
                    "assigned_to_name": assignee,
                    "assigned_to_email": assignee,
                }
            )

        result = import_team_records(items, SupportTicket, SupportTicket.Source.EXCEL, upsert=False)
        del request.session["support_import_headers"]
        del request.session["support_import_rows"]
        _flash_import_result(request, result)
        return redirect("support-analysis")

    mapping = {f["key"]: _guess_work_column(headers, f["key"]) for f in ANALYSIS_IMPORT_FIELDS}
    return render(
        request,
        "teams/support_import_map.html",
        {"headers": headers, "rows": rows[:8], "fields": ANALYSIS_IMPORT_FIELDS, "mapping": mapping},
    )
